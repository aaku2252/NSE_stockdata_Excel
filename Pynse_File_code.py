from pynse import *
import datetime 
from openpyxl import workbook,load_workbook
nse = Nse()

stocks = ['ASIANPAINT','AXISBANK','BHARTIARTL','COALINDIA','DLF','HINDALCO','INFY','SBIN','TATACONSUM','TATAMOTORS','TATASTEEL','TCS','TITAN','DEEPAKNTR','RELIANCE','HDFCBANK','VEDL','POLYCAB']

def stock_data(stocks,from_date:datetime.date= None):


    def populate_excel(symbol,from_date):

        wb = load_workbook(f'{symbol}.xlsx')
        ws = wb.active
        rows = ws.max_row
        last_date = from_date

        while rows>=1 and ws[f'a{rows}'].value==None and type(ws[f'a{rows}'].value)!=type(datetime.date.today()):
            rows-=1

        if rows>=2:
            last_date = ws[f'a{rows}'].value.date() or from_date

        trading_days = nse.get_hist(from_date=last_date,to_date=datetime.date.today()).index
        trading_days = list(trading_days.map(lambda x: x.date()))


        if rows>=2:
            if ws[f'a{rows}'].value.date()==trading_days[-1]:
                return
            
        

        def stock_analysis(symbol,from_date:datetime.date = None,to_date =trading_days[-1]):
            
            if rows>=2:
                trading_days.pop(0)
            if len(trading_days)>0:
                for date in trading_days:
                    bhav = nse.bhavcopy(date).loc[symbol].iloc[0]
                    bhav_fno_near = nse.bhavcopy_fno(date).loc[symbol].iloc[0]
                    bhav_fno_far = nse.bhavcopy_fno(date).loc[symbol].iloc[1]
                    hist = nse.get_hist(symbol=symbol,from_date=date,to_date=date)
                    vwap = bhav['CLOSE_PRICE']
                    if bhav['DATE1'] == trading_days[-1]:
                        vwap = nse.get_quote(symbol)['vwap'] 

                    data = [bhav['DATE1'],bhav['PREV_CLOSE'],bhav['OPEN_PRICE'],bhav['HIGH_PRICE'],bhav['LOW_PRICE'],bhav['CLOSE_PRICE'],vwap,bhav['TTL_TRD_QNTY'],float(bhav['DELIV_QTY']),float(bhav['DELIV_PER']),'',bhav['DATE1'],bhav_fno_near['EXPIRY_DT'].date(),bhav_fno_near['OPEN_INT'],bhav_fno_near['CHG_IN_OI'],'',bhav['DATE1'],bhav_fno_far['EXPIRY_DT'].date(),bhav_fno_far['OPEN_INT'],bhav_fno_far['CHG_IN_OI']]

                    ws.append(data)
                
        stock_analysis(symbol,last_date)
        print(symbol)
        wb.save(f'{symbol}.xlsx')
        wb.close()


    for symbol in stocks:
        populate_excel(symbol,from_date)


stock_data(stocks)
